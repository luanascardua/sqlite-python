from openpyxl import load_workbook
from datetime import datetime


class Sheet():


    def __init__(self, file = 'insertData.xlsx'):

        global sheets
        
        self.wb = load_workbook(file)
        self.file = file

        #armazena o nome de todas as sheets da pasta de trabalho do excel
        sheets = self.wb.sheetnames
    


    def countLines(self, sheet):

        global ws
    
        ws = self.wb[sheet]

        #armazena a última linha correspondente a coluna B
        self.lastLine = len(ws['B']) + 1

        data =  [
            x.value for x in ws['I']
            if x.value != None
        ]      
        #armazena a primeira linha vazia correspondente a coluna I
        self.firstLine = len(data) + 1


    def getDataUser(self, lines):
    
        self.name    = ws.cell(row=lines, column=2).value
        self.sigla   = ws.cell(row=lines, column=3).value
        self.idade   = ws.cell(row=lines, column=4).value

        
    def getDataTipoAvaliacao(self, lines):

        self.sigla     = ws.cell(row=lines, column=1).value
        self.descricao = ws.cell(row=lines, column=2).value


    def getDataAvaliacao(self, lines):

        self.sigla          = ws.cell(row=lines, column=2).value
        self.descricao      = ws.cell(row=lines, column=3).value
        self.tipoAvaliacao  = ws.cell(row=lines, column=4).value
        self.dataInicio     = ws.cell(row=lines, column=5).value
        self.dataTermino    = ws.cell(row=lines, column=6).value
        self.notaAprovacao  = ws.cell(row=lines, column=7).value
        self.idUser         = ws.cell(row=lines, column=8).value


    def getDataAvaliacaoAluno(self, lines):

        self.idUser         = ws.cell(row=lines, column=1).value
        self.idAvaliacao    = ws.cell(row=lines, column=2).value
        self.dataConclusao  = ws.cell(row=lines, column=3).value
        self.nota           = ws.cell(row=lines, column=4).value
        self.situacao       = ws.cell(row=lines, column=5).value


    def insertData(self, lines):

        if ws.cell(row=lines, column=9).value == None:
            ws.cell(row=lines, column=9).value = 'Inserido'
        
        if ws.cell(row=1, column=10).value == None:
            ws.cell(row=1, column=10).value = 'Horário'
        
        if ws.cell(row=lines, column=10).value == None:
            ws.cell(row=lines, column=10).value = datetime.now().strftime('%H:%M:%S %d/%m/%y')

        self.wb.save(self.file)


executeWb = Sheet()
#executeWb.countLines(sheets[0])
